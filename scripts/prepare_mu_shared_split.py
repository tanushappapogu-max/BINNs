"""Build a patient-level shared-training split from MU clinical metadata."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from gbm_pinn.mu_split import (
    assign_mu_roles,
    build_mu_shared_manifest,
    eligible_mu_patients,
)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--clinical-workbook", type=Path, required=True)
    parser.add_argument("--preserved-manifest", type=Path, required=True)
    parser.add_argument("--local-nifti-root", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--seed", type=int, default=162)
    args = parser.parse_args()
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ImportError("install the 'datasets' extra to read MU metadata") from error

    workbook = load_workbook(args.clinical_workbook, read_only=True, data_only=True)
    sheet = workbook["MU Glioma Post"]
    rows = sheet.iter_rows(values_only=True)
    headers = tuple(next(rows))
    patients = eligible_mu_patients(dict(zip(headers, row, strict=True)) for row in rows)
    preserved_value = json.loads(args.preserved_manifest.read_text(encoding="utf-8"))
    preserved_roles = {
        patient["patient_id"]: patient["role"] for patient in preserved_value["patients"]
    }
    assignments = assign_mu_roles(patients, preserved_roles=preserved_roles, seed=args.seed)
    manifest = build_mu_shared_manifest(
        patients,
        assignments,
        metadata_path=args.clinical_workbook,
        local_nifti_root=args.local_nifti_root,
        seed=args.seed,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )
    summary = {
        "eligible_patients": len(patients),
        "transitions": sum(patient.transition_count for patient in patients),
        "roles": {
            role: sum(value == role for value in assignments.values())
            for role in ("training", "model_selection", "final_test")
        },
        "locally_complete": sum(
            patient["local_images_complete"] for patient in manifest["patients"]
        ),
    }
    print(json.dumps(summary, indent=2, sort_keys=True))


if __name__ == "__main__":
    main()
